import os
import re
import time
import numpy as np
import open3d as o3d
import pandas as pd
from datetime import datetime
from sklearn.neighbors import NearestNeighbors
from openpyxl import Workbook, load_workbook  # ← 用於逐行寫入 Excel

# ------------------ 自訂體素下採樣函式 ------------------

def voxel_downsample(point_cloud, voxel_size):
    min_bounds = np.min(point_cloud, axis=0)
    voxel_indices = np.floor((point_cloud - min_bounds) / voxel_size).astype(np.int32)

    voxel_dict = {}
    for i, voxel_idx in enumerate(voxel_indices):
        key = tuple(voxel_idx)
        voxel_dict.setdefault(key, []).append(point_cloud[i])

    downsampled_points = []
    for points in voxel_dict.values():
        points = np.array(points)
        centroid = np.mean(points, axis=0)
        downsampled_points.append(centroid)

    return np.array(downsampled_points)


# ------------------ 自訂法向量計算函式 ------------------

def compute_normals(points, k=15, view_point=np.array([0, 0, 0], dtype=np.float64), tolerance=1e-8):
    n_points = points.shape[0]
    normals = np.zeros_like(points, dtype=np.float64)
    nbrs = NearestNeighbors(n_neighbors=k, algorithm='auto').fit(points)
    distances, indices = nbrs.kneighbors(points)

    for i in range(n_points):
        neighbor_pts = points[indices[i]]
        mean = neighbor_pts.mean(axis=0)
        cov = np.dot((neighbor_pts - mean).T, (neighbor_pts - mean)) / k
        eigenvalues, eigenvectors = np.linalg.eigh(cov)
        normal = eigenvectors[:, 0]

        norm_val = np.linalg.norm(normal)
        if norm_val > tolerance:
            normal = normal / norm_val
        else:
            normal = np.array([0, 0, 0], dtype=np.float64)

        if np.dot(points[i] - view_point, normal) < 0:
            normal = -normal

        normals[i] = normal

    return normals


# ------------------ 點雲預處理函式 ------------------

def custom_preprocess_point_cloud(point_cloud, voxel_size):
    points = np.asarray(point_cloud.points).astype(np.float64)
    down_points = voxel_downsample(points, voxel_size)
    normals = compute_normals(down_points, k=15, view_point=np.array([0, 0, 0], dtype=np.float64))

    down_pcd = o3d.geometry.PointCloud()
    down_pcd.points = o3d.utility.Vector3dVector(down_points)
    down_pcd.normals = o3d.utility.Vector3dVector(normals)
    return down_pcd


# ------------------ 配準與合併函式 ------------------

def refine_registration(source, target, initial_transformation, distance_threshold):
    result_icp = o3d.pipelines.registration.registration_icp(
        source, target,
        max_correspondence_distance=distance_threshold,
        init=initial_transformation,
        estimation_method=o3d.pipelines.registration.TransformationEstimationPointToPlane()
    )
    return result_icp


def process_pair(source_file, target_file, voxel_size, refined_distance_threshold, downsample=True):
    """
    處理單次配對，並回傳：
    - merged_pcd: 合併後的 PointCloud 物件
    - timing: 一個 dict，包含 IO、Preprocess、ICP、Write 各階段耗時
    """
    timing = {
        "Stage": None,
        "Source": os.path.basename(source_file),
        "Target": os.path.basename(target_file),
        "IO_time(s)": 0.0,
        "Preprocess_time(s)": 0.0,
        "ICP_time(s)": 0.0,
        "Write_time(s)": 0.0,
        "Total_time(s)": 0.0
    }

    # 1. IO 階段：讀取 Source 與 Target
    t_io_start = time.time()
    source_pcd = o3d.io.read_point_cloud(source_file)
    target_pcd = o3d.io.read_point_cloud(target_file)
    t_io_end = time.time()
    timing["IO_time(s)"] = round(t_io_end - t_io_start, 4)

    # 2. 若無法放下整雲則強制下採樣
    if not source_pcd.has_normals():
        source_pcd.estimate_normals(search_param=o3d.geometry.KDTreeSearchParamHybrid(radius=voxel_size * 2, max_nn=30))
    if not target_pcd.has_normals():
        target_pcd.estimate_normals(search_param=o3d.geometry.KDTreeSearchParamHybrid(radius=voxel_size * 2, max_nn=30))

    if not downsample:
        if len(source_pcd.points) > 500000 or len(target_pcd.points) > 500000:
            print("發現點雲點數超過 500,000，強制下採樣！")
            downsample = True

    # 3. Preprocess 階段：自訂下採樣與法向量
    t_pre_start = time.time()
    if downsample:
        source_proc = custom_preprocess_point_cloud(source_pcd, voxel_size)
        target_proc = custom_preprocess_point_cloud(target_pcd, voxel_size)
    else:
        source_proc = source_pcd
        target_proc = target_pcd

    if not source_proc.has_normals():
        source_proc.estimate_normals(search_param=o3d.geometry.KDTreeSearchParamHybrid(radius=voxel_size * 2, max_nn=30))
    if not target_proc.has_normals():
        target_proc.estimate_normals(search_param=o3d.geometry.KDTreeSearchParamHybrid(radius=voxel_size * 2, max_nn=30))
    t_pre_end = time.time()
    timing["Preprocess_time(s)"] = round(t_pre_end - t_pre_start, 4)

    # 4. ICP 階段
    initial_transformation = np.eye(4)
    t_icp_start = time.time()
    result_icp = refine_registration(source_proc, target_proc, initial_transformation, refined_distance_threshold)
    source_proc.transform(result_icp.transformation)
    t_icp_end = time.time()
    timing["ICP_time(s)"] = round(t_icp_end - t_icp_start, 4)

    # 5. Write 階段：將合併後的點雲作為 merged_pcd 返回（但實際寫入檔案的動作在主程式裡）
    t_write_start = time.time()
    merged_pcd = source_proc + target_proc
    merged_pcd.paint_uniform_color([1, 0, 0])
    t_write_end = time.time()
    timing["Write_time(s)"] = round(t_write_end - t_write_start, 4)

    # 計算 Total_time = 各階段之和
    timing["Total_time(s)"] = round(
        timing["IO_time(s)"] +
        timing["Preprocess_time(s)"] +
        timing["ICP_time(s)"] +
        timing["Write_time(s)"], 4
    )

    return merged_pcd, timing


def get_new_filename(file1, file2):
    pattern = r"(\d{5})_(\d{5})\.ply"
    m1 = re.match(pattern, file1)
    m2 = re.match(pattern, file2)
    if m1 and m2:
        left = m1.group(1)
        right = m2.group(2)
        return f"{left}_{right}.ply"
    else:
        return f"merged_{file1}_{file2}.ply"


# ------------------ Excel 逐行寫入輔助函式 ------------------

def append_to_excel(filepath, row_values, header=None):
    """
    如果檔案不存在，就建立新檔並寫入 header 和第一筆 row_values；
    否則直接打開並 append 一列 row_values。
    row_values: list，順序須對應 header。
    header: list of str，僅用於建立新檔時寫第一列欄位名稱。
    """
    if not os.path.exists(filepath):
        # 建立新 Workbook
        wb = Workbook()
        ws = wb.active
        if header is not None:
            ws.append(header)
        ws.append(row_values)
        wb.save(filepath)
    else:
        # 檔案已存在，直接讀取並 append
        wb = load_workbook(filepath)
        ws = wb.active
        ws.append(row_values)
        wb.save(filepath)


# ------------------ 主程式 ------------------

if __name__ == "__main__":
    print("配準開始時間：", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    voxel_size = 0.5
    refined_distance_threshold = voxel_size * 1.5
    base_folder = r"C:\Users\user\Desktop\PointCloud\red\furiren_ALL\icp_2"
    stage = 1
    current_folder = base_folder

    # Excel 相關參數
    excel_filename = "ICP_Timing_Log.xlsx"
    excel_path = os.path.join(base_folder, excel_filename)
    header = [
        "Stage", "Source", "Target",
        "IO_time(s)", "Preprocess_time(s)", "ICP_time(s)", "Write_time(s)", "Total_time(s)"
    ]

    # 刪除已存在的舊檔（如果想保留舊檔可註解此段）
    if os.path.exists(excel_path):
        os.remove(excel_path)

    while True:
        files = sorted([f for f in os.listdir(current_folder) if f.endswith(".ply")])
        num_files = len(files)
        print(f"\n=== Stage {stage}: {num_files} 檔案 ===")
        if num_files <= 1:
            print("只剩下一個檔案，結束合併！")
            break

        # 決定下採樣開關：stage=1 必下採樣，其餘每 5 階段下採樣
        ds_flag = (stage == 1) or (stage % 10 == 0)

        # 建立下一階段資料夾
        next_stage = stage + 1
        next_folder = os.path.join(base_folder, str(next_stage))
        os.makedirs(next_folder, exist_ok=True)

        for file1, file2 in zip(files[:-1], files[1:]):
            src_path = os.path.join(current_folder, file1)
            tgt_path = os.path.join(current_folder, file2)
            print(f"Stage {stage} 配準: {file1} + {file2} (downsample={ds_flag})")

            # 處理單次配對
            merged_pcd, timing = process_pair(src_path, tgt_path, voxel_size, refined_distance_threshold, downsample=ds_flag)

            # 記錄 Stage 編號
            timing["Stage"] = stage

            # 寫入檔案
            out_name = get_new_filename(file1, file2)
            out_path = os.path.join(next_folder, out_name)
            o3d.io.write_point_cloud(out_path, merged_pcd)
            print(f"→ 輸出: {out_name}，Total_time: {timing['Total_time(s)']} 秒")

            # 準備要寫入 Excel 的 row
            row = [
                timing["Stage"],
                timing["Source"],
                timing["Target"],
                timing["IO_time(s)"],
                timing["Preprocess_time(s)"],
                timing["ICP_time(s)"],
                timing["Write_time(s)"],
                timing["Total_time(s)"]
            ]

            # 每次配對完成就 append 到 Excel
            append_to_excel(excel_path, row_values=row, header=header)

        # 進入下一階段
        stage = next_stage
        current_folder = next_folder

    # 最後剩下一個檔案，印出提示
    final_files = [f for f in os.listdir(current_folder) if f.endswith(".ply")]
    if final_files:
        final_file = os.path.join(current_folder, final_files[0])
        print("最終成果檔案:", final_file)
    print("配準結束時間：", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    print(f"完整的耗時紀錄已逐次寫入：{excel_path}")
