import os
import re
import time
import shutil
import numpy as np
import open3d as o3d
import pandas as pd
from datetime import datetime
from sklearn.neighbors import NearestNeighbors
from openpyxl import Workbook, load_workbook

# ------------------ 自訂體素下採樣函式 ------------------

def voxel_downsample(point_cloud, voxel_size):
    min_bounds = np.min(point_cloud, axis=0)
    voxel_indices = np.floor((point_cloud - min_bounds) / voxel_size).astype(np.int32)

    voxel_dict = {}
    for i, voxel_idx in enumerate(voxel_indices):
        key = tuple(voxel_idx)
        voxel_dict.setdefault(key, []).append(point_cloud[i])

    downsampled_points = []
    for pts in voxel_dict.values():
        pts = np.array(pts)
        centroid = pts.mean(axis=0)
        downsampled_points.append(centroid)

    return np.array(downsampled_points)


# ------------------ 自訂法向量計算函式 ------------------

def compute_normals(points, k=15, view_point=np.array([0, 0, 0], dtype=np.float64), tolerance=1e-8):
    n_points = points.shape[0]
    normals = np.zeros_like(points, dtype=np.float64)
    nbrs = NearestNeighbors(n_neighbors=k, algorithm='auto').fit(points)
    _, indices = nbrs.kneighbors(points)

    for i in range(n_points):
        neighbor_pts = points[indices[i]]
        mu = neighbor_pts.mean(axis=0)
        cov = ((neighbor_pts - mu).T @ (neighbor_pts - mu)) / k
        _, eigvec = np.linalg.eigh(cov)
        normal = eigvec[:, 0]

        norm_val = np.linalg.norm(normal)
        if norm_val > tolerance:
            normal = normal / norm_val
        else:
            normal = np.zeros(3, dtype=np.float64)

        # 統一法向量方向向外
        if np.dot(points[i] - view_point, normal) < 0:
            normal = -normal

        normals[i] = normal

    return normals


# ------------------ 點雲預處理函式 ------------------

def custom_preprocess_point_cloud(pcd, voxel_size):
    points = np.asarray(pcd.points, dtype=np.float64)
    down_pts = voxel_downsample(points, voxel_size)
    normals = compute_normals(down_pts, k=15, view_point=np.array([0, 0, 0], dtype=np.float64))

    down_pcd = o3d.geometry.PointCloud()
    down_pcd.points = o3d.utility.Vector3dVector(down_pts)
    down_pcd.normals = o3d.utility.Vector3dVector(normals)
    return down_pcd


# ------------------ 配準與合併函式 ------------------

def refine_registration(src, tgt, init_trans, dist_thresh):
    result = o3d.pipelines.registration.registration_icp(
        src, tgt,
        max_correspondence_distance=dist_thresh,
        init=init_trans,
        estimation_method=o3d.pipelines.registration.TransformationEstimationPointToPlane()
    )
    return result


def process_pair(src_file, tgt_file, voxel_size, refined_thresh, downsample_flag=True):
    """
    處理單次配對，回傳：
      - merged_pcd（合併後點雲）
      - timing：一個 dict，包含 Stage、Source、Target、MergedPoints，
        IO_time(s)、Preprocess_time(s)、ICP_time(s)、Write_time(s)、Total_time(s)
    """
    timing = {
        "Stage": None,
        "Source": os.path.basename(src_file),
        "Target": os.path.basename(tgt_file),
        "MergedPoints": 0,
        "IO_time(s)": 0.0,
        "Preprocess_time(s)": 0.0,
        "ICP_time(s)": 0.0,
        "Write_time(s)": 0.0,
        "Total_time(s)": 0.0
    }

    # --- 1. IO 階段：讀檔 ---
    t_io0 = time.time()
    src_pcd = o3d.io.read_point_cloud(src_file)
    tgt_pcd = o3d.io.read_point_cloud(tgt_file)
    t_io1 = time.time()
    timing["IO_time(s)"] = round(t_io1 - t_io0, 4)

    # 若點數過多則強制下採樣
    if not src_pcd.has_normals():
        src_pcd.estimate_normals(search_param=o3d.geometry.KDTreeSearchParamHybrid(radius=voxel_size*2, max_nn=30))
    if not tgt_pcd.has_normals():
        tgt_pcd.estimate_normals(search_param=o3d.geometry.KDTreeSearchParamHybrid(radius=voxel_size*2, max_nn=30))
    if (not downsample_flag) and (len(src_pcd.points) > 500000 or len(tgt_pcd.points) > 500000):
        downsample_flag = True

    # --- 2. Preprocess 階段：下採樣＋重計算法向量 ---
    t_pre0 = time.time()
    if downsample_flag:
        src_proc = custom_preprocess_point_cloud(src_pcd, voxel_size)
        tgt_proc = custom_preprocess_point_cloud(tgt_pcd, voxel_size)
    else:
        src_proc = src_pcd
        tgt_proc = tgt_pcd
    # 確保有法向量
    if not src_proc.has_normals():
        src_proc.estimate_normals(search_param=o3d.geometry.KDTreeSearchParamHybrid(radius=voxel_size*2, max_nn=30))
    if not tgt_proc.has_normals():
        tgt_proc.estimate_normals(search_param=o3d.geometry.KDTreeSearchParamHybrid(radius=voxel_size*2, max_nn=30))
    t_pre1 = time.time()
    timing["Preprocess_time(s)"] = round(t_pre1 - t_pre0, 4)

    # --- 3. ICP 階段 ---
    init_trans = np.eye(4)
    t_icp0 = time.time()
    res_icp = refine_registration(src_proc, tgt_proc, init_trans, refined_thresh)
    src_proc.transform(res_icp.transformation)
    t_icp1 = time.time()
    timing["ICP_time(s)"] = round(t_icp1 - t_icp0, 4)

    # --- 4. Write 階段：合併但不寫檔（寫檔在外層） ---
    t_w0 = time.time()
    merged_pcd = src_proc + tgt_proc
    merged_pcd.paint_uniform_color([1, 0, 0])
    t_w1 = time.time()
    timing["Write_time(s)"] = round(t_w1 - t_w0, 4)

    # --- 5. 合併後點數記錄 ---
    timing["MergedPoints"] = int(np.asarray(merged_pcd.points).shape[0])

    # --- 6. 計算總耗時 ---
    total = (
        timing["IO_time(s)"] +
        timing["Preprocess_time(s)"] +
        timing["ICP_time(s)"] +
        timing["Write_time(s)"]
    )
    timing["Total_time(s)"] = round(total, 4)

    return merged_pcd, timing


def get_new_filename(f1, f2):
    pattern = r"(\d{5})_(\d{5})\.ply"
    m1 = re.match(pattern, f1)
    m2 = re.match(pattern, f2)
    if m1 and m2:
        return f"{m1.group(1)}_{m2.group(2)}.ply"
    else:
        return f"merged_{f1}_{f2}.ply"


# ------------------ Excel 逐行寫入輔助函式 ------------------

def append_to_excel(filepath, row_values, header=None):
    """
    如果檔案不存在，就建立新檔並寫入 header 和第一筆 row_values；
    否則直接讀取並在最底追加一列 row_values。
    """
    if not os.path.exists(filepath):
        wb = Workbook()
        ws = wb.active
        if header is not None:
            ws.append(header)
        ws.append(row_values)
        wb.save(filepath)
    else:
        wb = load_workbook(filepath)
        ws = wb.active
        ws.append(row_values)
        wb.save(filepath)


# ------------------ 主程式：連續執行 10 次 ------------------

if __name__ == "__main__":
    # 參數設定
    voxel_size = 0.5
    refined_dist_thresh = voxel_size * 1.5
    base_folder = r"C:\Users\user\Desktop\PointCloud\red\furiren_ALL\icp_2"

    # Excel 欄位標頭
    header = [
        "Stage", "Source", "Target", "MergedPoints",
        "IO_time(s)", "Preprocess_time(s)", "ICP_time(s)", "Write_time(s)", "Total_time(s)"
    ]

    # 執行 10 次
    for run_idx in range(1, 11):
        print(f"\n===== 第 {run_idx} 次執行（{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}）開始 =====")

        # 1. 每次執行前，先清空 base_folder 中所有子資料夾（stage_2, 3, …）：
        for name in os.listdir(base_folder):
            path = os.path.join(base_folder, name)
            if os.path.isdir(path):
                shutil.rmtree(path)

        # 2. 建立對應的 Excel 檔（若已存在就刪除）
        excel_filename = f"ICP_Timing_Log_run{run_idx}.xlsx"
        excel_path = os.path.join(base_folder, excel_filename)
        if os.path.exists(excel_path):
            os.remove(excel_path)

        # 3. 初始化 Stage 與 current_folder
        stage = 1
        current_folder = base_folder

        # 4. 開始二分階層合併
        while True:
            files = sorted([f for f in os.listdir(current_folder) if f.endswith(".ply")])
            num_files = len(files)
            print(f"  Stage {stage}: 找到 {num_files} 個 PLY 檔")
            if num_files <= 1:
                print(f"  Stage {stage} 結束（只剩最後一個檔案）")
                break

            # 決定是否下採樣：stage=1 一定下採樣，其餘每 5 階段下採樣
            ds_flag = (stage == 1) or (stage % 5 == 0)

            # 準備下一階段資料夾
            next_stage = stage + 1
            next_folder = os.path.join(base_folder, str(next_stage))
            os.makedirs(next_folder, exist_ok=True)

            # 逐對配對
            for f1, f2 in zip(files[:-1], files[1:]):
                src_path = os.path.join(current_folder, f1)
                tgt_path = os.path.join(current_folder, f2)
                print(f"    配對: {f1} + {f2} (downsample={ds_flag})")

                merged_pcd, timing = process_pair(src_path, tgt_path, voxel_size, refined_dist_thresh, downsample_flag=ds_flag)
                timing["Stage"] = stage

                # 寫出合併後點雲
                out_name = get_new_filename(f1, f2)
                out_path = os.path.join(next_folder, out_name)
                o3d.io.write_point_cloud(out_path, merged_pcd)
                print(f"      → 輸出: {out_name} (MergedPoints={timing['MergedPoints']})")

                # 準備要寫入 Excel 的一列（對應 header）
                row = [
                    timing["Stage"],
                    timing["Source"],
                    timing["Target"],
                    timing["MergedPoints"],
                    timing["IO_time(s)"],
                    timing["Preprocess_time(s)"],
                    timing["ICP_time(s)"],
                    timing["Write_time(s)"],
                    timing["Total_time(s)"]
                ]
                append_to_excel(excel_path, row_values=row, header=header)

            # 進入下一階段
            stage = next_stage
            current_folder = next_folder

        # 最後只剩一個檔案時，印出訊息
        final_files = [f for f in os.listdir(current_folder) if f.endswith(".ply")]
        if final_files:
            print(f"  [Run {run_idx}] 最終成果檔案: {os.path.join(current_folder, final_files[0])}")
        print(f"===== 第 {run_idx} 次執行結束（{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}），已輸出 {excel_filename} =====\n")

    print("所有 10 次執行皆已完成。")
